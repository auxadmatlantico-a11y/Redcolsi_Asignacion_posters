import logging
import re
import time
import unicodedata
from pathlib import Path
from typing import Optional, Tuple

from openpyxl import load_workbook
from playwright.sync_api import Playwright, sync_playwright, TimeoutError as PlaywrightTimeoutError


LOGIN_URL = "https://www.redcolsiplataforma.org/login.php?idPoison=ThePrincess6a020eb4e91039.280169151778519732"
DOCUMENTO = "1143453344"
CONTRASENA = "SemillaPCA2024"

# El archivo debe estar en la misma carpeta que este script.
EXCEL_INPUT = "REDCOLSI_2026_Asignacion.xlsx"
EXCEL_OUTPUT = "REDCOLSI_2026_Asignacion_resultado.xlsx"
SHEET_NAME = "Jornada_Miercoles_Mañana"

FECHA_FIJA = "2026-05-20"
HORA_INI = "10:00"
HORA_FIN = "13:00"
LUGAR = "Unilibre sede norte"
DESCRIPCION = "prueba"

LOG_FILE = "log.txt"


def setup_logging(base_dir: Path) -> None:
    log_path = base_dir / LOG_FILE
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )
    logging.info("Registro inicializado en %s", log_path)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


def extract_poster_number(value: object) -> str:
    """
    Convierte valores como P1, P02, 1, 02 a '1' o '2'.
    """
    if value is None:
        return ""
    text = str(value).strip()
    match = re.search(r"(\d+)", text)
    if match:
        return str(int(match.group(1)))
    return text


def load_projects(excel_path: Path) -> Tuple[list[dict], str, int]:
    wb = load_workbook(excel_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"No existe la hoja '{SHEET_NAME}'. Hojas disponibles: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    # Crear/ubicar columna de estado.
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    estado_col = None
    for idx, header in enumerate(headers, start=1):
        if normalize_text(header) == "ESTADO AUTOMATIZACION":
            estado_col = idx
            break

    if estado_col is None:
        estado_col = ws.max_column + 1
        ws.cell(1, estado_col).value = "Estado automatizacion"

    projects: list[dict] = []
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row, 2).value  # columna B
        poster = ws.cell(row, 10).value  # columna J
        if title is None or str(title).strip() == "":
            continue

        projects.append(
            {
                "row": row,
                "title": str(title).strip(),
                "poster": extract_poster_number(poster),
            }
        )

    return projects, excel_path.name, estado_col


def save_status(wb, output_path: Path) -> None:
    wb.save(output_path)


def safe_click(locator, timeout: int = 60000) -> None:
    locator.wait_for(state="visible", timeout=timeout)
    locator.scroll_into_view_if_needed(timeout=timeout)
    locator.click(timeout=timeout)


def set_date_with_js(page, date_value: str) -> None:
    """
    Intenta establecer la fecha directamente por JavaScript.
    Si el campo es readonly o la UI no permite escribir, esta vía suele funcionar.
    """
    locator = page.locator("#ProgramacionFecha")
    locator.wait_for(state="visible", timeout=60000)
    locator.scroll_into_view_if_needed(timeout=60000)

    # Abrir el selector de fecha.
    locator.click(timeout=60000)
    time.sleep(1)

    # Intento 1: fill normal
    try:
        locator.fill(date_value, timeout=5000)
    except Exception:
        # Intento 2: setter nativo + eventos
        locator.evaluate(
            """
            (el, value) => {
                const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
                setter.call(el, value);
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
                el.dispatchEvent(new Event('blur', { bubbles: true }));
            }
            """,
            date_value,
        )


def open_event_programming_page(page) -> None:
    page.goto(LOGIN_URL, wait_until="domcontentloaded")
    page.set_default_timeout(60000)

    page.get_by_role("textbox", name="Documento").wait_for(
        state="visible", timeout=60000)
    page.get_by_role("textbox", name="Documento").fill(DOCUMENTO)
    page.get_by_role("textbox", name="Contraseña").fill(CONTRASENA)
    safe_click(page.get_by_role("button", name="Inicia Sesión"))

    time.sleep(5)
    page.get_by_role("link", name=":: Ingresar ::.").nth(
        1).wait_for(state="visible", timeout=60000)
    safe_click(page.get_by_role("link", name=":: Ingresar ::.").nth(1))

    time.sleep(3)
    safe_click(page.get_by_role("link", name="   Eventos"))

    time.sleep(3)
    safe_click(page.get_by_role("link", name="2"))

    time.sleep(5)
    safe_click(page.locator(
        "tr:nth-child(11) > td:nth-child(8) > .dropdown > #dropdownMenuButton"))
    safe_click(page.get_by_role("link", name=re.compile(r"Editar", re.I)))

    time.sleep(5)
    safe_click(page.get_by_role("tab", name="Programación"))
    time.sleep(15)


def reset_current_form(page) -> None:
    try:
        reset_btn = page.get_by_role(
            "button", name=re.compile(r"^Reset$", re.I))
        if reset_btn.count() > 0:
            reset_btn.first.click(timeout=5000)
            time.sleep(2)
    except Exception:
        pass


def find_project_row(page, project_title: str):
    target = normalize_text(project_title)
    rows = page.locator("tr:has(.ProgramaChk)")
    count = rows.count()

    for i in range(count):
        row = rows.nth(i)
        try:
            row_text = normalize_text(row.inner_text(timeout=5000))
        except Exception:
            continue

        if target and target in row_text:
            return row

    return None


def assign_project(page, project_title: str, poster_number: str) -> None:
    # Fecha
    set_date_with_js(page, FECHA_FIJA)
    time.sleep(2)
    try:
        page.get_by_role("button", name="Ok").click(timeout=5000)
    except Exception:
        # En algunas pantallas el botón puede tardar un poco más en aparecer.
        page.get_by_role("button", name=re.compile(
            r"^Ok$", re.I)).click(timeout=10000)

    time.sleep(3)

    # Hora inicial y final
    ini = page.locator("#ProgramacionHoraIni")
    fin = page.locator("#ProgramacionHoraFin")

    ini.wait_for(state="visible", timeout=60000)
    fin.wait_for(state="visible", timeout=60000)

    ini.fill(HORA_INI)
    time.sleep(2)
    fin.fill(HORA_FIN)
    time.sleep(2)

    # Lugar y descripción
    page.get_by_role("textbox", name="Lugar programado").fill(LUGAR)
    time.sleep(2)
    page.get_by_role(
        "textbox", name="Descripción de la programación").fill(DESCRIPCION)
    time.sleep(2)

    # Poster
    poster_input = page.get_by_role("textbox", name="Rango inicial poster")
    poster_input.wait_for(state="visible", timeout=60000)
    poster_input.fill(poster_number)
    time.sleep(2)

    # Buscar el proyecto en la lista y marcar su checkbox.
    row = find_project_row(page, project_title)
    if row is None:
        raise ValueError(
            f"No se encontró el proyecto en la tabla: {project_title}")

    row.scroll_into_view_if_needed(timeout=60000)
    time.sleep(2)

    chk = row.locator(".ProgramaChk")
    chk.wait_for(state="visible", timeout=60000)
    if not chk.is_checked():
        chk.check(timeout=60000)
    time.sleep(2)

    # Guardar programación
    safe_click(page.get_by_role(
        "button", name=re.compile(r"Agregar programación", re.I)))
    time.sleep(3)

    # Aceptar confirmación si aparece
    try:
        safe_click(page.get_by_role(
            "button", name=re.compile(r"^Aceptar$", re.I)))
    except Exception:
        pass

    time.sleep(3)


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    setup_logging(base_dir)

    excel_path = base_dir / EXCEL_INPUT
    output_path = base_dir / EXCEL_OUTPUT

    if not excel_path.exists():
        raise FileNotFoundError(f"No se encontró el Excel: {excel_path}")

    # Cargar workbook una vez para actualizar estados.
    wb = load_workbook(excel_path)
    ws = wb[SHEET_NAME]

    # Crear columna de estado si no existe.
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    estado_col = None
    for idx, header in enumerate(headers, start=1):
        if normalize_text(header) == "ESTADO AUTOMATIZACION":
            estado_col = idx
            break
    if estado_col is None:
        estado_col = ws.max_column + 1
        ws.cell(1, estado_col).value = "Estado automatizacion"

    projects = []
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row, 2).value
        poster = ws.cell(row, 10).value
        if title is None or str(title).strip() == "":
            continue
        projects.append(
            {
                "row": row,
                "title": str(title).strip(),
                "poster": extract_poster_number(poster),
            }
        )

    if not projects:
        logging.warning("No hay proyectos para procesar.")
        wb.save(output_path)
        return

    logging.info("Se encontraron %d proyectos para procesar.", len(projects))

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=False,
            slow_mo=500,
            args=["--start-maximized"],
        )
        context = browser.new_context(no_viewport=True)
        page = context.new_page()
        page.set_default_timeout(60000)

        # Cerrar diálogos nativos si aparecieran.
        page.on("dialog", lambda dialog: dialog.accept())

        try:
            open_event_programming_page(page)

            # Limpia una vez al entrar.
            reset_current_form(page)

            for item in projects:
                row_num = item["row"]
                title = item["title"]
                poster = item["poster"] or ""

                logging.info(
                    "Procesando fila %s | Poster=%s | Título=%s", row_num, poster, title)
                try:
                    # Reabrir/asegurar el formulario en cada iteración.
                    # En la práctica ayuda cuando la plataforma conserva estado extraño.
                    page.get_by_role("tab", name="Programación").click(
                        timeout=60000)
                    time.sleep(3)

                    assign_project(page, title, poster)

                    ws.cell(row_num, estado_col).value = "OK"
                    save_status(wb, output_path)
                    logging.info("Fila %s completada correctamente.", row_num)

                    # Preparar la interfaz para la siguiente fila.
                    reset_current_form(page)

                except PlaywrightTimeoutError as e:
                    ws.cell(row_num, estado_col).value = "TIMEOUT"
                    save_status(wb, output_path)
                    logging.exception("TIMEOUT en fila %s: %s", row_num, e)
                    continue

                except Exception as e:
                    ws.cell(row_num, estado_col).value = "ERROR"
                    save_status(wb, output_path)
                    logging.exception("ERROR en fila %s: %s", row_num, e)
                    continue

        finally:
            try:
                save_status(wb, output_path)
            except Exception:
                pass
            context.close()
            browser.close()

    logging.info("Proceso finalizado. Archivo actualizado: %s", output_path)


if __name__ == "__main__":
    main()
